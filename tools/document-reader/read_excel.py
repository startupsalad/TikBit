#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel 读取 —— 抽成 markdown 表给 AI 读。

为什么默认抽文字而不是转图：Excel 是结构化数据，格子里的值就是核心，抽成表格
几乎不丢信息，比转图还好用（能搜、能算、不受分辨率限制）。这跟 PPT 正好相反。
只有带图表 / 复杂合并配色、需要看"长相"时，才用 --pdf 转 PDF 视觉读。

分工：
    .xlsx / .xlsm   openpyxl 直接读，不需要装 Office
    .xls            老二进制格式，走 pandas + xlrd
    .et             WPS 表格自家格式，走引擎转换
    --pdf           走引擎：Microsoft Office → WPS Office → LibreOffice 自动降级

用法：
    python read_excel.py <文件>              # 抽成 markdown 表（默认，最常用）
    python read_excel.py <文件> --sheet 名   # 只读指定工作表
    python read_excel.py <文件> --formula    # 显示公式而非计算值
    python read_excel.py <文件> --maxrows 50 # 每表最多输出多少行（大表防刷屏）
    python read_excel.py <文件> --pdf        # 转 PDF 视觉读（带图表时）
    python read_excel.py <文件> --out [路径]  # 存档到文件（默认只打印）

输出：默认只打印，不产生任何文件；--pdf 的产物落系统临时目录 docread/。
"""

import os
import sys
import warnings

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))  # 保证能 import engine
import engine

# openpyxl 对无默认样式的工作簿会刷 UserWarning，无害，压掉保持输出干净
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

DIRECT = ('.xlsx', '.xlsm')
EXTS = ('.xlsx', '.xlsm', '.xls', '.et')


def _cell(v):
    """格子值转成安全的表格单元：吃掉换行、转义竖线，免得把 markdown 表撑破"""
    if v is None:
        return ''
    if isinstance(v, float) and v != v:      # pandas NaN
        return ''
    return str(v).replace('\n', ' ').replace('|', '\\|').strip()


def _table(rows, maxrows=None):
    """把二维数据渲染成 markdown 表"""
    if not rows:
        return ['(空表)']
    ncol = max(len(r) for r in rows)
    out = []
    head = [_cell(c) for c in rows[0]] + [''] * (ncol - len(rows[0]))
    out.append('| ' + ' | '.join(head[:ncol]) + ' |')
    out.append('| ' + ' | '.join(['---'] * ncol) + ' |')
    body = rows[1:]
    shown = body[:maxrows] if maxrows else body
    for r in shown:
        cells = [_cell(c) for c in r] + [''] * (ncol - len(r))
        out.append('| ' + ' | '.join(cells[:ncol]) + ' |')
    if maxrows and len(body) > maxrows:
        out.append(f'…（还有 {len(body) - maxrows} 行，加 --maxrows 0 看全）')
    return out


def xlsx_to_text(path, only_sheet=None, show_formula=False, maxrows=None):
    """.xlsx/.xlsm → markdown 表（openpyxl，read_only 模式省内存）"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=not show_formula, read_only=True)
    out = []
    try:
        names = [only_sheet] if only_sheet else wb.sheetnames
        for name in names:
            if name not in wb.sheetnames:
                out.append(f'\n[未找到工作表: {name}]（本文件有: '
                           f'{"、".join(wb.sheetnames)}）\n')
                continue
            ws = wb[name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            while rows and all(c is None for c in rows[-1]):   # 去尾部全空行
                rows.pop()
            out.append(f'\n=== 工作表: {name}（{len(rows)} 行 × {ws.max_column} 列）===\n')
            out.extend(_table(rows, maxrows))
    finally:
        wb.close()
    return '\n'.join(out)


def xls_to_text(path, only_sheet=None, maxrows=None):
    """老 .xls → markdown 表（pandas + xlrd），按行原样 dump"""
    import pandas as pd
    xls = pd.ExcelFile(path)
    out = []
    names = [only_sheet] if only_sheet else xls.sheet_names
    for name in names:
        df = xls.parse(name, header=None)
        out.append(f'\n=== 工作表: {name}（{df.shape[0]} 行 × {df.shape[1]} 列）===\n')
        out.extend(_table(df.values.tolist(), maxrows))
    return '\n'.join(out)


def via_engine(path, only_sheet=None, maxrows=None):
    """.et 等格式：先用引擎转成 .xlsx 读不了，就转 PDF 抽文字兜底"""
    pdf, used = engine.office_to_pdf(path)
    print(f'[{os.path.splitext(path)[1]} 格式，经 {engine.ENGINE_NAME[used]} '
          f'转 PDF 后抽取，表格结构可能变形]', file=sys.stderr)
    import fitz
    with fitz.open(pdf) as d:
        return ''.join(f'\n=== 第 {i + 1} 页 ===\n' + p.get_text()
                       for i, p in enumerate(d))


def main():
    engine.fix_console()
    args = list(sys.argv[1:])
    want_pdf = engine.take_flag(args, '--pdf')
    show_formula = engine.take_flag(args, '--formula')
    only_sheet = engine.take_opt(args, '--sheet')
    maxrows = engine.take_opt(args, '--maxrows', int, 200)
    write_out = '--out' in args
    out_path = engine.take_opt(args, '--out')
    if maxrows == 0:
        maxrows = None          # --maxrows 0 = 不限行

    if not args:
        print('用法: python read_excel.py <文件> [--sheet 名] [--formula] '
              '[--maxrows 200] [--pdf] [--out [路径]]')
        sys.exit(1)

    src = args[0]
    if not os.path.exists(src):
        print(f'文件不存在: {src}')
        print('排查：微盘"按需下载"的占位符读不了，先在资源管理器里双击下载到本地。')
        sys.exit(1)

    ext = os.path.splitext(src)[1].lower()
    if ext not in EXTS:
        print(f'不支持的格式: {ext}（本工具处理 {"/".join(EXTS)}）')
        print('CSV 直接 Read 就行，不用这个脚本。')
        sys.exit(1)

    try:
        if want_pdf:
            pdf, used = engine.office_to_pdf(src, out=out_path)
            print(f'PDF 已生成（引擎 {engine.ENGINE_NAME[used]}）：\n{pdf}')
            print('→ AI 直接 Read 这个 PDF 即可视觉读取。')
            return
        if ext in DIRECT:
            content = xlsx_to_text(src, only_sheet, show_formula, maxrows)
        elif ext == '.xls':
            content = xls_to_text(src, only_sheet, maxrows)
        else:
            content = via_engine(src, only_sheet, maxrows)
    except Exception as exc:
        print(f'读取失败: {type(exc).__name__}: {exc}')
        print('排查：确认文件已下载到本地；--pdf 需要 Office / WPS / LibreOffice 任一。')
        sys.exit(1)

    if write_out:
        dst = out_path or (src + '.txt')
        with open(dst, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f'内容已保存到: {dst}')
    else:
        print(content)


if __name__ == '__main__':
    main()
